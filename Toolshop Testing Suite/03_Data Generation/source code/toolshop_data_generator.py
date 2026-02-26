import random
import json
import os

try:
    from faker import Faker
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: Missing libraries.")
    print("Please run: pip install faker openpyxl")
    exit()

# CONFIGURATION
fake = Faker()
Faker.seed(12345) 

# GLOBAL CACHES
category_cache = [] 
product_cache = []
user_cache = []

# CONSTANTS
CO2_RATINGS = [
    "None", "A (Lowest Impact)", "B (Low Impact)", 
    "C (Moderate Impact)", "D (Higher Impact)", "E (Highest Impact)"
]

TRANSACTION_STATUSES = [
    "AWAITING_FULFILLMENT", "ON_HOLD", 
    "AWAITING_SHIPMENT", "SHIPPED", "COMPLETED"
]

PAYMENT_METHODS = ["Cash on Delivery", "Credit Card", "Bank Transfer", "Gift Card", "Buy Now Pay Later"]

CATEGORY_STRUCTURE = {
    "Hand Tools": [
        "Hammer", "Claw Hammer", "Mallet", "Sledgehammer",
        "Wood Saw", "Hand Saw", "Hacksaw", "Chisel", "File",
        "Adjustable wrench", "Wrench", "Pipe Wrench", "Torque Wrench",
        "Open-end spanners (Set)", "Phillips Screwdriver", "Screwdriver", 
        "Pliers", "Combination Pliers", "Bolt Cutters", "Long Nose Pliers", "Slip Joint Pliers",
        "Utility Knife", "Tape Measure", "Level"
    ],
    "Power Tools": [
        "Sheet Sander", "Belt Sander", "Random Orbit Sander", "Sander",
        "Cordless Drill", "Cordless Drill 18V", "Drill Bits", "Drill",
        "Grinder", "Angle Grinder", "Circular Saw", "Jigsaw", "Reciprocating Saw", "Saw",
        "Heat Gun", "Router", "Planer"
    ],
    "Rentals": [
        "Crane", "Excavator", "Bulldozer", "Jackhammer", 
        "Concrete Mixer", "Generator", "Welding Machine", "Air Compressor", "Pressure Washer"
    ],
    "Other": [
        "Safety Goggles", "Work Gloves", "Tool Box", "Ladder", "Extension Cord", "Wheelbarrow"
    ]
}

# SAVE FUNCTION
def save_as_excel(filename, headers, data_rows):
    print(f"Saving {len(data_rows)} rows to {filename}...")
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        
    for row in data_rows:
        ws.append(row)
        
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(filename)

# GENERATION FUNCTIONS

def generate_categories(count=50):
    global category_cache
    headers = ["id", "parent_id", "name", "slug"]
    rows = []
    
    current_id = 1
    
    parents = {}
    for parent_name in CATEGORY_STRUCTURE.keys():
        slug = parent_name.lower().replace(" ", "-")
        rows.append([current_id, None, parent_name, slug]) 
        parents[parent_name] = current_id
        
        category_cache.append({"id": current_id, "name": parent_name})
        current_id += 1

    all_children = []
    for p_name, children in CATEGORY_STRUCTURE.items():
        for c_name in children:
            all_children.append((p_name, c_name))
            
    remaining_slots = count - len(parents)
    
    for i in range(remaining_slots):
        if i < len(all_children):
            parent_name, child_name = all_children[i]
            parent_id = parents[parent_name]
            name = child_name
        else:
            name = f"Specialty Tool {i}"
            parent_id = parents["Other"]

        slug = name.lower().replace(" ", "-")
        
        rows.append([current_id, parent_id, name, slug])
        
        # Cache for products
        category_cache.append({"id": current_id, "name": name})
        current_id += 1

    save_as_excel("categories.xlsx", headers, rows)

def generate_users(count=50):
    global user_cache 
    headers = ["id", "first_name", "last_name", "address", "city", "state", 
               "country", "postcode", "phone", "dob", "email", "password", "role"]
    rows = []
    
    for i in range(1, count + 1):
        fname = fake.first_name()
        lname = fake.last_name()
        addr = fake.street_address()
        city = fake.city()
        state = fake.state()
        country = fake.country()
        postcode = fake.postcode()
        dob_sql = fake.date_of_birth(minimum_age=18, maximum_age=70).strftime("%Y-%m-%d")
        
        clean_phone = fake.numerify('(###) ###-####')
        
        # Cache for transactions
        user_cache.append({
            "id": i,
            "name": f"{fname} {lname}",
            "address": addr,
            "city": city,
            "state": state,
            "country": country,
            "postcode": postcode
        })
        
        rows.append([
            i, fname, lname, addr, city, state, country, postcode,
            clean_phone, dob_sql, fake.email(), 
            "9e2ed9cb4bf54a6b9dc4669a1d295466b2585c4346092bffb5333098431cd61d", # Hash from SQL
            "user"
        ])

    save_as_excel("users.xlsx", headers, rows)

def generate_products(count=1000):
    global product_cache
    if not category_cache:
        print("Error: Category cache empty.")
        return

    headers = ["id", "name", "description", "stock", "price", "brand_id", 
               "category_id", "product_image_id", "is_location_offer", "is_rental", "co2_rating"]
    rows = []
    
    ADJECTIVES = ["durable", "lightweight", "heavy-duty", "ergonomic", "precision-engineered", 
                  "compact", "versatile", "high-performance", "reliable", "industry-standard"]
    
    USE_CASES = ["professional construction", "home DIY projects", "industrial applications", 
                 "precision tasks", "heavy lifting", "everyday repairs"]

    for i in range(1, count + 1):
        cat_data = random.choice(category_cache)
        cat_id = cat_data['id']
        cat_name = cat_data['name']
        
        name = f"{fake.color_name().capitalize()} {cat_name}"
        price = round(random.uniform(5.00, 200.00), 2)
        
        adj = random.choice(ADJECTIVES)
        use = random.choice(USE_CASES)
        description = f"This {adj} {name} is designed for {use}. It features a robust build quality ensuring long-lasting performance in any environment."

        mock_image_id = "01J" + fake.bothify(text='?#?#?#?#?#?#?#?#?#?#?#?#').upper()

        product_cache.append({
            "id": i,
            "name": name,
            "price": price
        })

        rows.append([
            i, name, description,
            random.randint(0, 100), 
            price,
            random.randint(1, 10), 
            cat_id,
            mock_image_id, 
            random.choice([0, 1]), 
            random.choice([0, 1]), 
            random.choice(CO2_RATINGS)
        ])

    save_as_excel("products.xlsx", headers, rows)

def generate_transactions(count=1000):
    if not product_cache or not user_cache:
        print("Error: Caches empty. Script execution order is wrong.")
        return

    headers = ["id", "user_id", "invoice_date", "invoice_number", 
               "billing_address", "billing_city", "billing_state", "billing_country", "billing_postcode",
               "total", "payment_method", "payment_account_name", "payment_account_number", 
               "created_at", "status", "purchased_items"]
    rows = []
    
    for i in range(1, count + 1):
        real_user = random.choice(user_cache)
        
        num_items = random.randint(1, 5)
        cart_items = random.sample(product_cache, num_items)
        cart_total = sum(item['price'] for item in cart_items)
        items_json = json.dumps(cart_items)
        
        inv_date = fake.date_time_between(start_date='-2y', end_date='now')
        inv_date_str = inv_date.strftime("%Y-%m-%d %H:%M:%S")

        rows.append([
            i,
            real_user['id'],
            inv_date_str,
            f"INV-{inv_date.year}{i:08d}",
            real_user['address'], 
            real_user['city'], 
            real_user['state'], 
            real_user['country'], 
            real_user['postcode'],
            round(cart_total, 2),
            random.choice(PAYMENT_METHODS),
            real_user['name'], 
            fake.bothify(text='#########???').upper(),
            inv_date_str,
            random.choice(TRANSACTION_STATUSES), 
            items_json                           
        ])

    save_as_excel("transactions.xlsx", headers, rows)

# MAIN EXECUTION
if __name__ == "__main__":
    generate_categories()
    generate_users()
    generate_products()
    generate_transactions()
    print("\nSUCCESS! 4 data files generated successfully.")